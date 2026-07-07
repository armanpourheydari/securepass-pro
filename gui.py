"""
gui.py - SecurePass Pro Complete GUI
"""

import math
import sys
from datetime import datetime

import pyperclip
from openpyxl import Workbook
from PyQt6.QtCore import (
    QEasingCurve,
    QEvent,
    QPropertyAnimation,
    Qt,
    QTimer,
    pyqtProperty,
)
from PyQt6.QtGui import QColor, QKeyEvent, QPalette
from PyQt6.QtWidgets import (
    QApplication,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSizePolicy,
    QSlider,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from controller import PasswordController


class AnimatedProgressBar(QProgressBar):
    """Custom progress bar with smooth animation and dynamic color changes."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._target_value = 0
        self._current_value = 0
        self.animation = QPropertyAnimation(self, b"animatedValue")
        self.animation.setDuration(600)
        self.animation.setEasingCurve(QEasingCurve.Type.OutCubic)
        self.animation.valueChanged.connect(self._update_value)
        self.setRange(0, 100)

    @pyqtProperty(int)
    def animatedValue(self):
        return self._current_value

    @animatedValue.setter
    def animatedValue(self, value):
        self._current_value = value
        super().setValue(value)
        self._update_color(value)

    def setValue(self, value):
        self.animation.stop()
        self.animation.setStartValue(self._current_value)
        self.animation.setEndValue(value)
        self.animation.start()

    def _update_value(self, value):
        self._current_value = value
        super().setValue(value)
        self._update_color(value)

    def _update_color(self, value):
        if value >= 85:
            color = "#4a9eff"
        elif value >= 70:
            color = "#4caf50"
        elif value >= 50:
            color = "#ffc107"
        elif value >= 30:
            color = "#ff9800"
        else:
            color = "#f44336"
        self.setStyleSheet(f"""
            QProgressBar::chunk {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {color}, stop:1 {color}cc);
                border-radius: 6px;
            }}
        """)


class DeselectableTable(QTableWidget):
    """QTableWidget that clears selection when clicking on empty area."""

    def mousePressEvent(self, event):
        if self.itemAt(event.position().toPoint()) is None:
            self.clearSelection()
        super().mousePressEvent(event)


class SavePasswordDialog(QDialog):
    """Dialog for saving a new password with show/hide functionality."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Save New Password")
        self.setMinimumSize(380, 220)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(14)

        layout.addWidget(QLabel("Name:"))
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("e.g., Gmail, GitHub")
        layout.addWidget(self.name_input)

        layout.addWidget(QLabel("Password:"))
        pwd_layout = QHBoxLayout()
        pwd_layout.setSpacing(8)

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Enter password")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        pwd_layout.addWidget(self.password_input)

        self.eye_btn = QPushButton("👁️")
        self.eye_btn.setFixedSize(36, 36)
        self.eye_btn.setStyleSheet("""
            QPushButton {
                background-color: #2a2a48;
                color: #a0a0b8;
                border: 1px solid #2a2a48;
                border-radius: 8px;
                font-size: 14px;
                padding: 0px;
                margin: 0px;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #3a3a5a;
                border-color: #4a6fa5;
            }
        """)
        self.eye_btn.clicked.connect(self._toggle_password_visibility)
        pwd_layout.addWidget(self.eye_btn)

        layout.addLayout(pwd_layout)

        # Enter key navigation
        self.name_input.returnPressed.connect(self.password_input.setFocus)
        self.password_input.returnPressed.connect(self.accept)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("Save")
        save_btn.clicked.connect(self.accept)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setObjectName("cancel")
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #0f0f1a, stop:1 #1a1a2e);
                border-radius: 12px;
            }
            QLabel {
                color: #e8e8f0;
                font-family: 'Segoe UI';
                font-size: 12px;
                font-weight: 600;
            }
            QLineEdit {
                background-color: #14142a;
                color: #f2c94c;
                border: 1px solid #2a2a48;
                border-radius: 8px;
                padding: 10px 14px;
                font-size: 13px;
                font-family: 'Segoe UI';
            }
            QLineEdit:focus {
                border-color: #4a6fa5;
            }
            QPushButton {
                background-color: #4a6fa5;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 8px 20px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #5a7fb5;
            }
            QPushButton#cancel {
                background-color: #2a2a48;
            }
            QPushButton#cancel:hover {
                background-color: #3a3a5a;
            }
        """)

    def _toggle_password_visibility(self):
        if self.password_input.echoMode() == QLineEdit.EchoMode.Password:
            self.password_input.setEchoMode(QLineEdit.EchoMode.Normal)
            self.eye_btn.setText("🙈")
        else:
            self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
            self.eye_btn.setText("👁️")

    def get_data(self):
        return (self.name_input.text().strip(), self.password_input.text().strip())


class CardDialog(QDialog):
    """Dialog for adding/editing a credit/debit card."""

    def __init__(self, parent=None, card_data=None):
        super().__init__(parent)
        self.card_data = card_data
        self.setWindowTitle("Add Card" if not card_data else "Edit Card")
        self.setMinimumSize(360, 450)
        self.setup_ui()
        if card_data:
            self._populate_fields()

    def setup_ui(self):
        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #0f0f1a, stop:1 #1a1a2e);
                border-radius: 16px;
            }
            QLabel {
                color: #e8e8f0;
                font-family: 'Segoe UI';
                font-size: 13px;
                font-weight: 600;
            }
            QLineEdit, QComboBox {
                background-color: #14142a;
                color: #f2c94c;
                border: 1px solid #2a2a48;
                border-radius: 10px;
                padding: 10px 14px;
                font-size: 13px;
                font-family: 'Segoe UI';
            }
            QLineEdit:focus, QComboBox:focus {
                border-color: #4a6fa5;
            }
            QComboBox QAbstractItemView {
                background-color: #14142a;
                color: #e8e8f0;
                border: 1px solid #2a2a48;
                selection-background-color: #4a6fa5;
            }
            QPushButton {
                background-color: #4a6fa5;
                color: white;
                border: none;
                border-radius: 10px;
                padding: 10px 24px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #5a7fb5;
            }
            QPushButton#cancel {
                background-color: #2a2a48;
            }
            QPushButton#cancel:hover {
                background-color: #3a3a5a;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setSpacing(14)

        # ========== فیلدها ==========
        layout.addWidget(QLabel("Card Name:"))
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Personal, Business, etc.")
        layout.addWidget(self.name_input)

        layout.addWidget(QLabel("Card Type:"))
        self.type_combo = QComboBox()
        self.type_combo.addItems(["visa", "mastercard", "amex", "discover", "other"])
        layout.addWidget(self.type_combo)

        layout.addWidget(QLabel("Card Number:"))
        self.number_input = QLineEdit()
        self.number_input.setPlaceholderText("1234 5678 9012 3456")
        layout.addWidget(self.number_input)

        layout.addWidget(QLabel("Card Holder:"))
        self.holder_input = QLineEdit()
        self.holder_input.setPlaceholderText("John Doe")
        layout.addWidget(self.holder_input)

        expiry_layout = QHBoxLayout()
        expiry_layout.addWidget(QLabel("Expiry (MM/YY):"))

        self.expiry_month = QLineEdit()
        self.expiry_month.setPlaceholderText("MM")
        self.expiry_month.setFixedWidth(50)
        self.expiry_month.setMaxLength(2)
        self.expiry_month.textChanged.connect(
            lambda t: self._validate_digit(t, self.expiry_month)
        )
        expiry_layout.addWidget(self.expiry_month)

        self.expiry_year = QLineEdit()
        self.expiry_year.setPlaceholderText("YY")
        self.expiry_year.setFixedWidth(50)
        self.expiry_year.setMaxLength(2)
        self.expiry_year.textChanged.connect(
            lambda t: self._validate_digit(t, self.expiry_year)
        )
        expiry_layout.addWidget(self.expiry_year)

        expiry_layout.addStretch()
        layout.addLayout(expiry_layout)

        layout.addWidget(QLabel("CVV:"))
        self.cvv_input = QLineEdit()
        self.cvv_input.setPlaceholderText("123")
        self.cvv_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.cvv_input.setMaxLength(4)
        self.cvv_input.textChanged.connect(
            lambda t: self._validate_digit(t, self.cvv_input)
        )
        layout.addWidget(self.cvv_input)

        # ========== دکمه‌ها ==========
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        self.save_btn = QPushButton("Save Card")
        self.save_btn.setAutoDefault(False)  # ✅ جلوگیری از پیش‌فرض شدن
        self.save_btn.clicked.connect(self.accept)
        btn_layout.addWidget(self.save_btn)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("cancel")
        cancel_btn.setAutoDefault(False)  # ✅ جلوگیری از پیش‌فرض شدن
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

        # ========== Enter Key Navigation (بعد از دکمه‌ها) ==========
        # از Name مستقیماً به Number (ComboBox را رد می‌کند)
        self.name_input.returnPressed.connect(self.number_input.setFocus)
        self.number_input.returnPressed.connect(self.holder_input.setFocus)
        self.holder_input.returnPressed.connect(self.expiry_month.setFocus)
        self.expiry_month.returnPressed.connect(self.expiry_year.setFocus)
        self.expiry_year.returnPressed.connect(self.cvv_input.setFocus)
        # در CVV با Enter ثبت می‌شود
        self.cvv_input.returnPressed.connect(self.save_btn.click)

    def _validate_digit(self, text, field):
        if text and not text.isdigit():
            field.setText("".join(filter(str.isdigit, text)))

    def _populate_fields(self):
        self.name_input.setText(self.card_data.get("name", ""))
        self.type_combo.setCurrentText(self.card_data.get("type", "visa"))
        self.number_input.setText(self.card_data.get("number", ""))
        self.holder_input.setText(self.card_data.get("holder", ""))
        self.expiry_month.setText(self.card_data.get("expiry_month", ""))
        self.expiry_year.setText(self.card_data.get("expiry_year", ""))
        self.cvv_input.setText(self.card_data.get("cvv", ""))

    def get_data(self) -> dict:
        data = {
            "name": self.name_input.text().strip(),
            "type": self.type_combo.currentText(),
            "number": self.number_input.text().strip().replace(" ", ""),
            "holder": self.holder_input.text().strip(),
            "expiry_month": self.expiry_month.text().strip(),
            "expiry_year": self.expiry_year.text().strip(),
            "cvv": self.cvv_input.text().strip(),
        }

        errors = []
        if not data["name"]:
            errors.append("Card Name")
        if not data["number"] or len(data["number"]) < 4:
            errors.append("Card Number (at least 4 digits)")
        if not data["holder"]:
            errors.append("Card Holder")
        if not data["expiry_month"] or not data["expiry_year"]:
            errors.append("Expiry Date")
        if not data["cvv"]:
            errors.append("CVV")

        if errors:
            raise ValueError(f"Please fill in: {', '.join(errors)}")

        if not data["number"].isdigit():
            raise ValueError("Card number must contain only digits")
        if not data["cvv"].isdigit():
            raise ValueError("CVV must contain only digits")
        if not data["expiry_month"].isdigit() or not data["expiry_year"].isdigit():
            raise ValueError("Expiry must be numbers")

        return data


class NoteDialog(QDialog):
    """Dialog for adding/editing a secure note."""

    def __init__(self, parent=None, note_data=None):
        super().__init__(parent)
        self.note_data = note_data
        self.setWindowTitle("New Note" if not note_data else "Edit Note")
        self.setMinimumSize(420, 380)
        self.setup_ui()
        if note_data:
            self._populate_fields()

    def setup_ui(self):
        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #0f0f1a, stop:1 #1a1a2e);
                border-radius: 16px;
            }
            QLabel {
                color: #e8e8f0;
                font-family: 'Segoe UI';
                font-size: 13px;
                font-weight: 600;
            }
            QLineEdit, QTextEdit {
                background-color: #14142a;
                color: #e8e8f0;
                border: 1px solid #2a2a48;
                border-radius: 10px;
                padding: 10px 14px;
                font-size: 13px;
                font-family: 'Segoe UI';
            }
            QLineEdit:focus, QTextEdit:focus {
                border-color: #4a6fa5;
            }
            QPushButton {
                background-color: #4a6fa5;
                color: white;
                border: none;
                border-radius: 10px;
                padding: 10px 24px;
                font-weight: bold;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: #5a7fb5;
            }
            QPushButton#cancel {
                background-color: #2a2a48;
            }
            QPushButton#cancel:hover {
                background-color: #3a3a5a;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setSpacing(14)

        # ========== Title ==========
        layout.addWidget(QLabel("Title:"))
        self.title_input = QLineEdit()
        self.title_input.setPlaceholderText("Enter note title...")
        layout.addWidget(self.title_input)

        # ========== Content ==========
        layout.addWidget(QLabel("Content:"))
        self.content_input = QTextEdit()
        self.content_input.setPlaceholderText("Write your secure note here...")
        self.content_input.setMinimumHeight(160)
        self.content_input.installEventFilter(self)  # برای گرفتن Ctrl+Enter
        layout.addWidget(self.content_input)

        # ========== Buttons ==========
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        self.save_btn = QPushButton("Save Note")
        self.save_btn.setAutoDefault(False)  # ✅ جلوگیری از پیش‌فرض شدن
        self.save_btn.clicked.connect(self.accept)
        btn_layout.addWidget(self.save_btn)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setObjectName("cancel")
        cancel_btn.setAutoDefault(False)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

        # ========== Enter Key Navigation ==========
        self.title_input.returnPressed.connect(self.content_input.setFocus)

    def eventFilter(self, obj, event):
        """گرفتن Ctrl+Enter در QTextEdit برای ثبت"""
        if obj == self.content_input and event.type() == QEvent.Type.KeyPress:
            if (
                event.key() == Qt.Key.Key_Return
                and event.modifiers() == Qt.KeyboardModifier.ControlModifier
            ):
                self.save_btn.click()
                return True
        return super().eventFilter(obj, event)

    def _populate_fields(self):
        self.title_input.setText(self.note_data.get("title", ""))
        self.content_input.setPlainText(self.note_data.get("content", ""))

    def get_data(self) -> dict:
        return {
            "title": self.title_input.text().strip(),
            "content": self.content_input.toPlainText().strip(),
        }


class SecurePassGUI(QMainWindow):
    """Main application window with real-time password analysis and vault."""

    def __init__(self):
        super().__init__()
        self.controller = PasswordController()
        self.master_password = None
        self.current_password = ""

        self._init_widgets()
        self.setup_window()
        self.setup_ui()
        self.center_window()

    def _init_widgets(self):
        """Initialize all widget references."""
        self.length_slider = None
        self.length_display = None
        self.password_display = None
        self.check_input = None
        self.check_eye_btn = None
        self.result_card = None
        self.score_label = None
        self.level_label = None
        self.progress_bar = None
        self.suggestions_text = None
        self.copy_btn = None
        self.strength_bar = None
        self.strength_label = None
        self.vault_table = None
        self.cards_table = None
        self.notes_table = None
        self.vault_content_stack = None
        self.vault_sub_tabs = None
        self.master_entry = None
        self.show_hide_btn = None
        self.entropy_label = None
        self.crack_time_label = None
        self.gen_entropy_label = None
        self.gen_crack_label = None
        self.gen_strength_label = None
        self.gen_progress = None
        self.detail_items = []
        self.tabs = None

    def setup_window(self):
        self.setWindowTitle("SecurePass Pro")
        self.setMinimumSize(520, 700)
        self.setMaximumSize(660, 860)
        self.setStyleSheet(self._get_style_sheet())

    def center_window(self):
        screen = QApplication.primaryScreen().availableGeometry()
        width = int(screen.width() * 0.38)
        height = int(width * 16 / 9)
        if height > screen.height() * 0.88:
            height = int(screen.height() * 0.88)
            width = int(height * 9 / 16)
        x = (screen.width() - width) // 2
        y = (screen.height() - height) // 2
        self.setGeometry(x, y, width, height)

    def _get_style_sheet(self):
        return """
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #0a0a1a, stop:1 #15152a);
            }
            QLabel {
                color: #e8e8f0;
                font-family: 'Segoe UI';
            }
            QLabel#title {
                color: #f2c94c;
                font-size: 28px;
                font-weight: 800;
                letter-spacing: 1.5px;
            }
            QLabel#subtitle {
                color: #8888aa;
                font-size: 13px;
                letter-spacing: 0.5px;
            }
            QLabel#section_title {
                color: #d4d4e8;
                font-size: 14px;
                font-weight: 700;
                letter-spacing: 0.3px;
                margin-bottom: 4px;
            }
            QLabel#score_label {
                font-size: 44px;
                font-weight: 800;
            }
            QLabel#level_label {
                font-size: 16px;
                font-weight: 700;
            }
            QLabel#detail_label {
                color: #a0a0b8;
                font-size: 11px;
                font-weight: 500;
            }
            QLabel#detail_value {
                color: #e8e8f0;
                font-size: 11px;
                font-weight: 600;
            }
            QLineEdit {
                background-color: #14142a;
                color: #f2c94c;
                border: 2px solid #2a2a48;
                border-radius: 12px;
                padding: 14px 18px;
                font-size: 15px;
                font-family: 'Consolas', monospace;
                font-weight: 600;
            }
            QLineEdit:focus {
                border-color: #4a6fa5;
            }
            QLineEdit#check_input {
                color: #e8e8f0;
                font-family: 'Segoe UI';
                font-weight: 500;
                font-size: 15px;
                padding: 16px 22px;
                background-color: #14142a;
                border: 2px solid #2a2a48;
                border-radius: 14px;
            }
            QLineEdit#check_input:focus {
                border-color: #4a6fa5;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #4a6fa5, stop:1 #3a5f95);
                color: white;
                border: none;
                border-radius: 12px;
                font-family: 'Segoe UI';
                font-size: 14px;
                font-weight: 700;
                padding: 14px 28px;
                letter-spacing: 0.5px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #5a7fb5, stop:1 #4a6fa5);
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3a5f95, stop:1 #2a4f85);
            }
            QPushButton#generate_btn {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #4a8a5a, stop:1 #3a7a4a);
            }
            QPushButton#generate_btn:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #5a9a6a, stop:1 #4a8a5a);
            }
            QPushButton#copy_btn {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2a2a48, stop:1 #1a1a36);
                border: 1px solid #3a3a5a;
                font-size: 12px;
                padding: 10px 18px;
                border-radius: 10px;
            }
            QPushButton#copy_btn:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3a3a5a, stop:1 #2a2a48);
                border-color: #4a6fa5;
            }
            QPushButton#export_btn {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2a5a3a, stop:1 #1a4a2a);
                font-size: 11px;
                padding: 8px 14px;
                border-radius: 8px;
            }
            QPushButton#export_btn:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3a6a4a, stop:1 #2a5a3a);
            }
            QPushButton#delete_btn {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #a54a4a, stop:1 #8a3a3a);
            }
            QPushButton#delete_btn:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #b55a5a, stop:1 #9a4a4a);
            }
            QSlider::groove:horizontal {
                height: 6px;
                background: #2a2a48;
                border-radius: 3px;
            }
            QSlider::handle:horizontal {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #f2c94c, stop:1 #f5a623);
                width: 24px;
                height: 24px;
                margin: -9px 0;
                border-radius: 12px;
            }
            QSlider::handle:horizontal:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #f5d76e, stop:1 #f2c94c);
            }
            QProgressBar {
                border: none;
                background-color: #1a1a2e;
                border-radius: 8px;
                height: 12px;
            }
            QProgressBar::chunk {
                border-radius: 8px;
            }
            QTableWidget {
                background-color: #14142a;
                color: #e8e8f0;
                border: 1px solid #2a2a48;
                border-radius: 10px;
                gridline-color: #2a2a48;
            }
            QTableWidget::item {
                padding: 8px;
            }
            QTableWidget::item:selected {
                background-color: #4a6fa5;
                color: white;
            }
            QHeaderView::section {
                background-color: #1a1a2e;
                color: #f2c94c;
                padding: 10px;
                border: none;
                font-weight: bold;
            }
            QTabWidget::pane {
                border: none;
                background: transparent;
            }
            QTabBar::tab {
                background-color: #1a1a2e;
                color: #8888aa;
                padding: 12px 28px;
                border-top-left-radius: 12px;
                border-top-right-radius: 12px;
                font-family: 'Segoe UI';
                font-size: 14px;
                font-weight: 700;
                letter-spacing: 0.5px;
            }
            QTabBar::tab:selected {
                background-color: #2a2a48;
                color: #f2c94c;
            }
            QTabBar::tab:hover:!selected {
                background-color: #22223a;
                color: #b0b0c8;
            }
            QFrame#glass_card {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(26, 26, 46, 0.75), stop:1 rgba(20, 20, 40, 0.60));
                border: 1px solid rgba(255, 255, 255, 0.08);
                border-radius: 20px;
                padding: 20px;
            }
            QFrame#glass_card_dark {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(20, 20, 40, 0.85), stop:1 rgba(12, 12, 30, 0.75));
                border: 1px solid rgba(74, 111, 165, 0.15);
                border-radius: 20px;
                padding: 20px;
            }
        """

    def setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(24, 20, 24, 20)
        main_layout.setSpacing(16)

        self._build_header(main_layout)

        self.tabs = QTabWidget()
        self.tabs.addTab(self._build_checker_tab(), "🔍 Checker")
        self.tabs.addTab(self._build_generator_tab(), "⚡ Generator")

        vault_tab = QWidget()
        vault_layout = QVBoxLayout(vault_tab)
        vault_layout.setContentsMargins(0, 0, 0, 0)
        self.vault_sub_tabs = QTabWidget()
        self.vault_sub_tabs.setVisible(False)
        vault_layout.addWidget(self.vault_sub_tabs)
        self.tabs.addTab(vault_tab, "📂 Vault")

        self.tabs.currentChanged.connect(self._on_tab_changed)
        main_layout.addWidget(self.tabs)

        self._build_footer(main_layout)
        self._build_vault_tab()

    def _build_header(self, parent):
        header = QVBoxLayout()
        header.setSpacing(4)

        title = QLabel("🔐 SecurePass Pro")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.addWidget(title)

        subtitle = QLabel("Advanced Password Security Suite")
        subtitle.setObjectName("subtitle")
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.addWidget(subtitle)

        parent.addLayout(header)

    def _build_checker_tab(self):
        """Build the real-time password checker tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(16)

        # Input card
        input_card = QFrame()
        input_card.setObjectName("glass_card")
        input_layout = QVBoxLayout(input_card)
        input_layout.setSpacing(12)

        input_label = QLabel("🔑 Enter Password to Check")
        input_label.setObjectName("section_title")
        input_layout.addWidget(input_label)

        # Input with eye button
        input_container = QHBoxLayout()
        input_container.setSpacing(8)

        self.check_input = QLineEdit()
        self.check_input.setObjectName("check_input")
        self.check_input.setPlaceholderText("Type or paste your password here...")
        self.check_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.check_input.textChanged.connect(self._on_real_time_check)
        input_container.addWidget(self.check_input)

        self.check_eye_btn = QPushButton("👁️")
        self.check_eye_btn.setFixedSize(50, 50)
        self.check_eye_btn.setStyleSheet("""
            QPushButton {
                background-color: #2a2a48;
                color: #a0a0b8;
                border-radius: 10px;
                font-size: 18px;
                padding: 0px;
                margin: 0px;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #3a3a5a;
                border-color: #4a6fa5;
            }
        """)
        self.check_eye_btn.setToolTip("Show/Hide password")
        self.check_eye_btn.clicked.connect(self._toggle_check_password_visibility)
        input_container.addWidget(self.check_eye_btn)

        input_layout.addLayout(input_container)

        layout.addWidget(input_card)

        # Results card
        self.result_card = QFrame()
        self.result_card.setObjectName("glass_card_dark")
        self.result_card.setVisible(False)
        result_layout = QVBoxLayout(self.result_card)
        result_layout.setSpacing(14)

        # Score and level
        score_layout = QHBoxLayout()
        score_layout.setSpacing(20)

        left_score = QVBoxLayout()
        left_score.setSpacing(0)
        self.score_label = QLabel("--")
        self.score_label.setObjectName("score_label")
        self.score_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        left_score.addWidget(self.score_label)

        self.level_label = QLabel("--")
        self.level_label.setObjectName("level_label")
        self.level_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        left_score.addWidget(self.level_label)

        score_layout.addLayout(left_score, 1)

        right_progress = QVBoxLayout()
        right_progress.setSpacing(8)
        self.progress_bar = AnimatedProgressBar()
        self.progress_bar.setValue(0)
        right_progress.addWidget(self.progress_bar)

        info_layout = QHBoxLayout()
        info_layout.setSpacing(12)
        self.entropy_label = QLabel("🔢 Entropy: -- bits")
        self.entropy_label.setObjectName("detail_label")
        self.entropy_label.setWordWrap(True)
        info_layout.addWidget(self.entropy_label)

        self.crack_time_label = QLabel("⏱️ Crack time: --")
        self.crack_time_label.setObjectName("detail_label")
        self.crack_time_label.setWordWrap(True)
        info_layout.addWidget(self.crack_time_label)

        info_layout.addStretch()
        right_progress.addLayout(info_layout)

        score_layout.addLayout(right_progress, 3)
        result_layout.addLayout(score_layout)

        # Suggestions
        suggestions_label = QLabel("💡 Suggestions")
        suggestions_label.setObjectName("section_title")
        suggestions_label.setStyleSheet("margin-top: 4px;")
        result_layout.addWidget(suggestions_label)

        self.suggestions_text = QLabel("")
        self.suggestions_text.setWordWrap(True)
        self.suggestions_text.setStyleSheet("""
            color: #a0a0b8;
            font-size: 13px;
            font-family: 'Segoe UI';
            padding: 4px 0px;
        """)
        result_layout.addWidget(self.suggestions_text)

        generate_btn = QPushButton("🚀 Generate Strong Password")
        generate_btn.setObjectName("generate_btn")
        generate_btn.clicked.connect(self._go_to_generator)
        result_layout.addWidget(generate_btn)

        layout.addWidget(self.result_card)
        layout.addStretch()
        return tab

    def _toggle_check_password_visibility(self):
        if self.check_input.echoMode() == QLineEdit.EchoMode.Password:
            self.check_input.setEchoMode(QLineEdit.EchoMode.Normal)
            self.check_eye_btn.setText("🙈")
            self.check_eye_btn.setToolTip("Hide password")
        else:
            self.check_input.setEchoMode(QLineEdit.EchoMode.Password)
            self.check_eye_btn.setText("👁️")
            self.check_eye_btn.setToolTip("Show password")

    def _build_generator_tab(self):
        """Build the password generator tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(16)

        # Generator card
        gen_card = QFrame()
        gen_card.setObjectName("glass_card")
        gen_layout = QVBoxLayout(gen_card)
        gen_layout.setSpacing(12)

        length_label = QLabel("📏 Password Length")
        length_label.setObjectName("section_title")
        gen_layout.addWidget(length_label)

        slider_layout = QHBoxLayout()
        slider_layout.setSpacing(14)

        self.length_slider = QSlider(Qt.Orientation.Horizontal)
        self.length_slider.setMinimum(6)
        self.length_slider.setMaximum(40)
        self.length_slider.setValue(16)
        self.length_slider.valueChanged.connect(self._update_length_display)

        self.length_display = QLabel("16")
        self.length_display.setStyleSheet("""
            font-size: 22px;
            font-weight: 800;
            color: #f2c94c;
            min-width: 44px;
            font-family: 'Inter';
        """)
        self.length_display.setAlignment(Qt.AlignmentFlag.AlignCenter)

        slider_layout.addWidget(self.length_slider)
        slider_layout.addWidget(self.length_display)
        gen_layout.addLayout(slider_layout)

        generate_btn = QPushButton("🔑 Generate Secure Password")
        generate_btn.clicked.connect(self._on_generate)
        gen_layout.addWidget(generate_btn)

        self.password_display = QLineEdit()
        self.password_display.setReadOnly(True)
        self.password_display.setPlaceholderText(
            "Your secure password will appear here"
        )
        gen_layout.addWidget(self.password_display)

        # Copy button with reference
        self.copy_btn = QPushButton("📋 Copy to Clipboard")
        self.copy_btn.setObjectName("copy_btn")
        self.copy_btn.clicked.connect(self._copy_password)
        gen_layout.addWidget(self.copy_btn)

        layout.addWidget(gen_card)

        # Strength card
        strength_card = QFrame()
        strength_card.setObjectName("glass_card_dark")
        strength_layout = QVBoxLayout(strength_card)
        strength_layout.setSpacing(10)

        strength_label = QLabel("💪 Password Strength")
        strength_label.setObjectName("section_title")
        strength_layout.addWidget(strength_label)

        self.gen_progress = AnimatedProgressBar()
        self.gen_progress.setValue(0)
        strength_layout.addWidget(self.gen_progress)

        self.gen_strength_label = QLabel("")
        self.gen_strength_label.setStyleSheet("font-size: 16px; font-weight: 700;")
        self.gen_strength_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        strength_layout.addWidget(self.gen_strength_label)

        gen_info_layout = QHBoxLayout()
        gen_info_layout.setSpacing(20)

        self.gen_entropy_label = QLabel("🔢 Entropy: -- bits")
        self.gen_entropy_label.setObjectName("detail_label")
        gen_info_layout.addWidget(self.gen_entropy_label)

        self.gen_crack_label = QLabel("⏱️ Crack time: --")
        self.gen_crack_label.setObjectName("detail_label")
        gen_info_layout.addWidget(self.gen_crack_label)

        gen_info_layout.addStretch()
        strength_layout.addLayout(gen_info_layout)

        layout.addWidget(strength_card)
        layout.addStretch()
        return tab

    def _build_vault_tab(self):
        """Build the vault tab with unlock screen."""
        tab = QWidget()
        vault_layout = QVBoxLayout(tab)
        vault_layout.setSpacing(0)

        self.vault_content_stack = QStackedWidget()

        # Unlock screen
        unlock_widget = QWidget()
        unlock_layout = QVBoxLayout(unlock_widget)
        unlock_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        unlock_layout.setSpacing(16)

        title = QLabel("🔐 Secure Vault")
        title.setStyleSheet("""
            font-size: 32px;
            font-weight: 800;
            color: #f2c94c;
            font-family: 'Segoe UI';
        """)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        unlock_layout.addWidget(title)

        subtitle = QLabel("Enter your master password to access saved data")
        subtitle.setStyleSheet("""
            font-size: 14px;
            color: #8888aa;
            font-family: 'Segoe UI';
        """)
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        unlock_layout.addWidget(subtitle)

        unlock_layout.addSpacing(10)

        self.master_entry = QLineEdit()
        self.master_entry.setPlaceholderText("Master Password")
        self.master_entry.setEchoMode(QLineEdit.EchoMode.Password)
        self.master_entry.setStyleSheet("""
            QLineEdit {
                background-color: #14142a;
                color: #f2c94c;
                border: 2px solid #2a2a48;
                border-radius: 12px;
                padding: 16px 22px;
                font-size: 16px;
                font-family: 'Segoe UI';
                max-width: 300px;
                min-height: 48px;
            }
            QLineEdit:focus {
                border-color: #4a6fa5;
            }
        """)
        self.master_entry.setMaximumWidth(300)
        self.master_entry.setMinimumHeight(48)
        self.master_entry.returnPressed.connect(self._unlock_vault_action)
        unlock_layout.addWidget(
            self.master_entry, alignment=Qt.AlignmentFlag.AlignCenter
        )

        self.show_hide_btn = QPushButton("👁️")
        self.show_hide_btn.setFixedSize(48, 48)
        self.show_hide_btn.setStyleSheet("""
            QPushButton {
                background-color: #2a2a48;
                color: #a0a0b8;
                border: 2px solid #2a2a48;
                border-radius: 12px;
                font-size: 18px;
                padding: 0px;
                margin: 0px;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #3a3a5a;
                border-color: #4a6fa5;
            }
        """)
        self.show_hide_btn.setToolTip("Show/Hide password")
        self.show_hide_btn.clicked.connect(self._toggle_master_password_visibility)
        master_layout = QHBoxLayout()
        master_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        master_layout.setSpacing(8)
        master_layout.addWidget(self.master_entry)
        master_layout.addWidget(self.show_hide_btn)
        unlock_layout.addLayout(master_layout)

        unlock_btn = QPushButton("🔓 Unlock Vault")
        unlock_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #4a6fa5, stop:1 #3a5f95);
                color: white;
                border: none;
                border-radius: 12px;
                font-family: 'Segoe UI';
                font-size: 16px;
                font-weight: 700;
                padding: 16px 36px;
                max-width: 300px;
                min-height: 48px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #5a7fb5, stop:1 #4a6fa5);
            }
        """)
        unlock_btn.setMaximumWidth(300)
        unlock_btn.setMinimumHeight(48)
        unlock_btn.clicked.connect(self._unlock_vault_action)
        unlock_layout.addWidget(unlock_btn, alignment=Qt.AlignmentFlag.AlignCenter)

        unlock_layout.addStretch()

        # Vault content
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setSpacing(10)

        self.vault_sub_tabs = QTabWidget()
        self.vault_sub_tabs.addTab(self._build_passwords_tab(), "🔑 Passwords")
        self.vault_sub_tabs.addTab(self._build_cards_tab(), "💳 Cards")
        self.vault_sub_tabs.addTab(self._build_notes_tab(), "📝 Notes")
        self.vault_sub_tabs.setVisible(False)
        content_layout.addWidget(self.vault_sub_tabs)

        self.vault_content_stack.addWidget(unlock_widget)
        self.vault_content_stack.addWidget(content_widget)

        vault_layout.addWidget(self.vault_content_stack)

        vault_tab_index = self.tabs.indexOf(self.tabs.widget(2))
        if vault_tab_index >= 0:
            self.tabs.widget(vault_tab_index).layout().addWidget(tab)

    def _build_passwords_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(10)

        header = QHBoxLayout()
        title = QLabel("Saved Passwords")
        title.setObjectName("section_title")
        header.addWidget(title)
        header.addStretch()

        save_btn = QPushButton("➕ Add New")
        save_btn.setObjectName("export_btn")
        save_btn.clicked.connect(self._save_new_password)
        header.addWidget(save_btn)

        refresh_btn = QPushButton("🔄 Refresh")
        refresh_btn.setObjectName("export_btn")
        refresh_btn.clicked.connect(self._load_vault)
        header.addWidget(refresh_btn)

        layout.addLayout(header)

        self.vault_table = DeselectableTable()
        self.vault_table.setColumnCount(2)
        self.vault_table.setHorizontalHeaderLabels(["Name", "Password"])
        self.vault_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents
        )
        self.vault_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        self.vault_table.setAlternatingRowColors(True)
        self.vault_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.vault_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.vault_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.vault_table.cellClicked.connect(self._on_password_cell_clicked)
        self.vault_table.cellDoubleClicked.connect(self._edit_password)
        layout.addWidget(self.vault_table)

        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(8)

        self.delete_btn = QPushButton("🗑️ Delete")
        self.delete_btn.setObjectName("delete_btn")
        self.delete_btn.clicked.connect(self._delete_selected_password)
        btn_layout.addWidget(self.delete_btn)

        btn_layout.addStretch()

        self.export_passwords_txt_btn = QPushButton("📄 Export TXT")
        self.export_passwords_txt_btn.setObjectName("export_btn")
        self.export_passwords_txt_btn.clicked.connect(self._export_passwords_txt)
        btn_layout.addWidget(self.export_passwords_txt_btn)

        self.export_passwords_excel_btn = QPushButton("📊 Export Excel")
        self.export_passwords_excel_btn.setObjectName("export_btn")
        self.export_passwords_excel_btn.clicked.connect(self._export_passwords_excel)
        btn_layout.addWidget(self.export_passwords_excel_btn)

        layout.addLayout(btn_layout)

        return tab

    def _build_cards_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(10)

        header = QHBoxLayout()
        title = QLabel("Saved Cards")
        title.setObjectName("section_title")
        header.addWidget(title)
        header.addStretch()

        add_card_btn = QPushButton("➕ Add New")
        add_card_btn.setObjectName("export_btn")
        add_card_btn.clicked.connect(self._add_new_card)
        header.addWidget(add_card_btn)

        refresh_cards_btn = QPushButton("🔄 Refresh")
        refresh_cards_btn.setObjectName("export_btn")
        refresh_cards_btn.clicked.connect(self._load_cards)
        header.addWidget(refresh_cards_btn)

        layout.addLayout(header)

        self.cards_table = DeselectableTable()
        self.cards_table.setColumnCount(5)
        self.cards_table.setHorizontalHeaderLabels(
            ["Name", "Type", "Number", "Expiry", "Holder"]
        )
        self.cards_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self.cards_table.setAlternatingRowColors(True)
        self.cards_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.cards_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.cards_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.cards_table.cellDoubleClicked.connect(self._edit_card)
        layout.addWidget(self.cards_table)

        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(8)

        self.delete_btn_card = QPushButton("🗑️ Delete")
        self.delete_btn_card.setObjectName("delete_btn")
        self.delete_btn_card.clicked.connect(self._delete_selected_card)
        btn_layout.addWidget(self.delete_btn_card)

        btn_layout.addStretch()

        self.export_cards_txt_btn = QPushButton("📄 Export TXT")
        self.export_cards_txt_btn.setObjectName("export_btn")
        self.export_cards_txt_btn.clicked.connect(self._export_cards_txt)
        btn_layout.addWidget(self.export_cards_txt_btn)

        self.export_cards_excel_btn = QPushButton("📊 Export Excel")
        self.export_cards_excel_btn.setObjectName("export_btn")
        self.export_cards_excel_btn.clicked.connect(self._export_cards_excel)
        btn_layout.addWidget(self.export_cards_excel_btn)

        layout.addLayout(btn_layout)

        return tab

    def _build_notes_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(10)

        header = QHBoxLayout()
        title = QLabel("Secure Notes")
        title.setObjectName("section_title")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #f2c94c;")
        header.addWidget(title)
        header.addStretch()

        add_note_btn = QPushButton("➕ Add New")
        add_note_btn.setObjectName("export_btn")
        add_note_btn.clicked.connect(self._add_new_note)
        header.addWidget(add_note_btn)

        refresh_notes_btn = QPushButton("🔄 Refresh")
        refresh_notes_btn.setObjectName("export_btn")
        refresh_notes_btn.clicked.connect(self._load_notes)
        header.addWidget(refresh_notes_btn)

        layout.addLayout(header)

        self.notes_table = DeselectableTable()
        self.notes_table.setColumnCount(3)
        self.notes_table.setHorizontalHeaderLabels(["Title", "Content", "View"])
        self.notes_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents
        )
        self.notes_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        self.notes_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Fixed
        )
        self.notes_table.setColumnWidth(2, 80)
        self.notes_table.setAlternatingRowColors(True)
        self.notes_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.notes_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.notes_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.notes_table.setStyleSheet("""
            QTableWidget {
                background-color: #14142a;
                color: #e8e8f0;
                border: 1px solid #2a2a48;
                border-radius: 10px;
                gridline-color: #2a2a48;
            }
            QTableWidget::item {
                padding: 8px;
            }
            QTableWidget::item:selected {
                background-color: #4a6fa5;
                color: white;
            }
            QHeaderView::section {
                background-color: #1a1a2e;
                color: #f2c94c;
                padding: 10px;
                border: none;
                font-weight: bold;
            }
        """)
        self.notes_table.cellDoubleClicked.connect(self._edit_note)
        layout.addWidget(self.notes_table)

        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(8)

        self.delete_note_btn = QPushButton("🗑️ Delete")
        self.delete_note_btn.setObjectName("delete_btn")
        self.delete_note_btn.clicked.connect(self._delete_selected_note)
        btn_layout.addWidget(self.delete_note_btn)

        btn_layout.addStretch()

        self.export_notes_txt_btn = QPushButton("📄 Export TXT")
        self.export_notes_txt_btn.setObjectName("export_btn")
        self.export_notes_txt_btn.clicked.connect(self._export_notes_txt)
        btn_layout.addWidget(self.export_notes_txt_btn)

        self.export_notes_excel_btn = QPushButton("📊 Export Excel")
        self.export_notes_excel_btn.setObjectName("export_btn")
        self.export_notes_excel_btn.clicked.connect(self._export_notes_excel)
        btn_layout.addWidget(self.export_notes_excel_btn)

        layout.addLayout(btn_layout)

        return tab

    def _build_footer(self, parent):
        footer = QLabel(
            "🔒 All operations are performed locally. Your passwords never leave your device."
        )
        footer.setStyleSheet("color: #606080; font-size: 9px; font-family: 'Segoe UI';")
        footer.setAlignment(Qt.AlignmentFlag.AlignCenter)
        parent.addWidget(footer)

    def _toggle_master_password_visibility(self):
        if self.master_entry.echoMode() == QLineEdit.EchoMode.Password:
            self.master_entry.setEchoMode(QLineEdit.EchoMode.Normal)
            self.show_hide_btn.setText("🙈")
            self.show_hide_btn.setToolTip("Hide password")
        else:
            self.master_entry.setEchoMode(QLineEdit.EchoMode.Password)
            self.show_hide_btn.setText("👁️")
            self.show_hide_btn.setToolTip("Show password")

    def _on_tab_changed(self, index):
        if self.tabs.tabText(index) == "Vault":
            self.master_password = None
            self.controller.set_master_password(None)
            self.vault_content_stack.setCurrentIndex(0)
            self.vault_sub_tabs.setVisible(False)
            self.master_entry.clear()
            self.master_entry.setFocus()

    def _update_length_display(self, value):
        self.length_display.setText(str(value))

    def _on_generate(self):
        length = self.length_slider.value()
        password = self.controller.generate(length)
        self.password_display.setText(password)
        self._update_gen_strength(password)

    def _update_gen_strength(self, password):
        result = self.controller.check(password)
        score = result["score"]

        self.gen_progress.setValue(score)

        if score >= 85:
            color = "#4a9eff"
            level_text = "🔵 Excellent"
        elif score >= 70:
            color = "#4caf50"
            level_text = "🟢 Good"
        elif score >= 50:
            color = "#ffc107"
            level_text = "🟡 Medium"
        elif score >= 30:
            color = "#ff9800"
            level_text = "🟠 Weak"
        else:
            color = "#f44336"
            level_text = "🔴 Very Weak"

        self.gen_progress.setStyleSheet(f"""
            QProgressBar::chunk {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {color}, stop:1 {color}cc);
                border-radius: 6px;
            }}
        """)
        self.gen_strength_label.setText(f"{level_text} ({score}%)")
        self.gen_strength_label.setStyleSheet(
            f"color: {color}; font-size: 16px; font-weight: 700;"
        )

        entropy = result.get("entropy", 0)
        crack_time = result.get("crack_time", {}).get("time_str", "--")
        self.gen_entropy_label.setText(f"🔢 Entropy:\n{entropy:.2f} bits")
        self.gen_crack_label.setText(f"⏱️ Crack time:\n{crack_time}")

    def _copy_password(self):
        """Copy generated password to clipboard."""
        pwd = self.password_display.text()
        if pwd:
            pyperclip.copy(pwd)
            if self.copy_btn:
                self.copy_btn.setText("✅ Copied!")
                self.copy_btn.setStyleSheet("""
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #4a8a5a, stop:1 #3a7a4a);
                    color: white;
                    border: none;
                    border-radius: 6px;
                    font-family: 'Segoe UI';
                    font-size: 11px;
                    padding: 8px 12px;
                    font-weight: bold;
                """)
                QTimer.singleShot(1500, self._reset_copy_btn)
        else:
            QMessageBox.warning(self, "Empty", "Generate a password first!")

    def _reset_copy_btn(self):
        if self.copy_btn:
            self.copy_btn.setText("📋 Copy to Clipboard")
            self.copy_btn.setStyleSheet("")

    def _on_real_time_check(self):
        """Real-time password analysis triggered on every text change."""
        password = self.check_input.text()
        self.current_password = password

        if len(password) < 3:
            self.result_card.setVisible(False)
            return

        self.result_card.setVisible(True)

        result = self.controller.check(password)
        score = result["score"]
        level = result["level"]

        self.score_label.setText(f"{score}%")
        self.progress_bar.setValue(score)

        if score >= 85:
            color = "#4a9eff"
        elif score >= 70:
            color = "#4caf50"
        elif score >= 50:
            color = "#ffc107"
        elif score >= 30:
            color = "#ff9800"
        else:
            color = "#f44336"

        self.level_label.setText(level)
        self.level_label.setStyleSheet(
            f"color: {color}; font-size: 16px; font-weight: 700;"
        )

        entropy = result.get("entropy", 0)
        crack_time = result.get("crack_time", {}).get("time_str", "--")
        self.entropy_label.setText(f"🔢 Entropy:\n{entropy:.2f} bits")
        self.crack_time_label.setText(f"⏱️ Crack time:\n{crack_time}")

        self._update_details(result)

        suggestions = result.get("suggestions", [])
        if suggestions:
            self.suggestions_text.setText("\n".join(f"• {s}" for s in suggestions[:3]))
            self.suggestions_text.setStyleSheet("""
                color: #f28b82;
                font-size: 13px;
                font-family: 'Segoe UI';
            """)
        else:
            self.suggestions_text.setText("✅ No suggestions. Password looks great!")
            self.suggestions_text.setStyleSheet("""
                color: #a0d8a0;
                font-size: 13px;
                font-family: 'Segoe UI';
            """)

    def _update_details(self, result):
        detail_values = {
            "length_val": f"{len(self.current_password)} chars",
            "upper_val": "✅" if result.get("has_upper", False) else "❌",
            "digit_val": "✅" if result.get("has_digit", False) else "❌",
            "symbol_val": "✅" if result.get("has_symbol", False) else "❌",
            "common_val": "⚠️" if result.get("is_common", False) else "✅",
            "variety_val": f"{result.get('variety', 0)}/4",
        }

        for key, item_layout in self.detail_items:
            for child in item_layout.children():
                if isinstance(child, QLabel) and child.objectName() == "detail_value":
                    if key in detail_values:
                        child.setText(detail_values[key])

    def _go_to_generator(self):
        self.tabs.setCurrentIndex(1)

    def _unlock_vault_action(self):
        master = self.master_entry.text()
        if not master:
            QMessageBox.warning(self, "Empty", "Please enter your master password.")
            return

        if self.controller.verify_password(master):
            self.master_password = master
            self.controller.set_master_password(master)
            self.vault_sub_tabs.setVisible(True)
            self.vault_content_stack.setCurrentIndex(1)
            self._load_vault()
            self._load_cards()
            self._load_notes()
            self.master_entry.clear()
        else:
            if not self.controller.vault_exists():
                reply = QMessageBox.question(
                    self,
                    "Create New Vault",
                    "No vault found. Create a new vault with this master password?\n\n"
                    "⚠️ Make sure you remember this password!",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                )
                if reply == QMessageBox.StandardButton.Yes:
                    if self.controller.create_new_vault(master):
                        self.master_password = master
                        self.controller.set_master_password(master)
                        self.vault_sub_tabs.setVisible(True)
                        self.vault_content_stack.setCurrentIndex(1)
                        self._load_vault()
                        self._load_cards()
                        self._load_notes()
                        self.master_entry.clear()
                        QMessageBox.information(
                            self, "Success", "Vault created successfully!"
                        )
                    else:
                        QMessageBox.warning(self, "Error", "Failed to create vault.")
            else:
                QMessageBox.warning(self, "Error", "Wrong master password!")

    def _save_new_password(self):
        if not self.master_password:
            return

        dialog = SavePasswordDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            name, password = dialog.get_data()
            if not name or not password:
                QMessageBox.warning(self, "Error", "Name and password are required!")
                return

            try:
                self.controller.set_master_password(self.master_password)
                if self.controller.save_password(name, password):
                    QMessageBox.information(
                        self, "Success", f"Password '{name}' saved!"
                    )
                    self._load_vault()
                else:
                    QMessageBox.warning(self, "Error", "Failed to save password.")
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))

    def _load_vault(self):
        if not self.master_password:
            return

        try:
            self.controller.set_master_password(self.master_password)
            passwords = self.controller.get_passwords()
            self.vault_table.setRowCount(len(passwords))
            for row, pwd in enumerate(passwords):
                self.vault_table.setItem(row, 0, QTableWidgetItem(pwd.get("name", "")))
                self.vault_table.setItem(
                    row, 1, QTableWidgetItem(pwd.get("password", ""))
                )
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load vault: {str(e)}")

    def _on_password_cell_clicked(self, row, column):
        if column == 1:
            item = self.vault_table.item(row, column)
            if item:
                pyperclip.copy(item.text())
                QMessageBox.information(self, "Copied", "Password copied to clipboard!")

    def _edit_password(self, row, column):
        name_item = self.vault_table.item(row, 0)
        pass_item = self.vault_table.item(row, 1)
        if not name_item or not pass_item:
            return

        current_name = name_item.text()
        current_password = pass_item.text()

        new_name, ok = QInputDialog.getText(
            self, "Edit Password", "Name:", text=current_name
        )
        if not ok:
            return

        new_password, ok = QInputDialog.getText(
            self,
            "Edit Password",
            "Password:",
            QLineEdit.EchoMode.Password,
            text=current_password,
        )
        if not ok:
            return

        try:
            self.controller.set_master_password(self.master_password)
            if self.controller.update_password(row + 1, new_name, new_password):
                QMessageBox.information(self, "Success", "Password updated!")
                self._load_vault()
            else:
                QMessageBox.warning(self, "Error", "Failed to update password.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _delete_selected_password(self):
        selected = self.vault_table.currentRow()
        if selected < 0:
            QMessageBox.warning(
                self, "No Selection", "Please select a password to delete."
            )
            return

        name_item = self.vault_table.item(selected, 0)
        name = name_item.text() if name_item else "Unknown"

        reply = QMessageBox.question(
            self,
            "Delete Password",
            f"Are you sure you want to delete '{name}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.controller.set_master_password(self.master_password)
                if self.controller.delete_password(selected + 1):
                    QMessageBox.information(self, "Success", "Password deleted!")
                    self._load_vault()
                else:
                    QMessageBox.warning(self, "Error", "Failed to delete password.")
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))

    def _add_new_card(self):
        if not self.master_password:
            QMessageBox.warning(self, "Locked", "Please unlock vault first!")
            return

        dialog = CardDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            try:
                data = dialog.get_data()
            except ValueError as e:
                QMessageBox.warning(self, "Invalid Input", str(e))
                return

            try:
                self.controller.set_master_password(self.master_password)
                if self.controller.save_card(
                    data["name"],
                    data["number"],
                    data["holder"],
                    data["expiry_month"],
                    data["expiry_year"],
                    data["cvv"],
                    data["type"],
                ):
                    QMessageBox.information(self, "Success", "Card saved successfully!")
                    self._load_cards()
                else:
                    QMessageBox.warning(self, "Error", "Failed to save card.")
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))

    def _load_cards(self):
        if not self.master_password:
            return

        try:
            self.controller.set_master_password(self.master_password)
            cards = self.controller.get_cards()
            self.cards_table.setRowCount(len(cards))
            for row, card in enumerate(cards):
                self.cards_table.setItem(row, 0, QTableWidgetItem(card.get("name", "")))
                self.cards_table.setItem(row, 1, QTableWidgetItem(card.get("type", "")))
                number = card.get("number", "")
                masked = "•••• " + number[-4:] if len(number) >= 4 else "••••"
                self.cards_table.setItem(row, 2, QTableWidgetItem(masked))
                expiry = f"{card.get('expiry_month', '')}/{card.get('expiry_year', '')}"
                self.cards_table.setItem(row, 3, QTableWidgetItem(expiry))
                self.cards_table.setItem(
                    row, 4, QTableWidgetItem(card.get("holder", ""))
                )
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load cards: {str(e)}")

    def _edit_card(self, row, column):
        if not self.master_password:
            return

        try:
            self.controller.set_master_password(self.master_password)
            cards = self.controller.get_cards()
            if row >= len(cards):
                return

            dialog = CardDialog(self, cards[row])
            if dialog.exec() == QDialog.DialogCode.Accepted:
                data = dialog.get_data()
                if not data["name"] or not data["number"] or not data["holder"]:
                    QMessageBox.warning(
                        self, "Error", "Name, number and holder are required!"
                    )
                    return
                if not data["number"].isdigit() or not data["cvv"].isdigit():
                    QMessageBox.warning(
                        self, "Error", "Card number and CVV must contain only digits!"
                    )
                    return

                if self.controller.update_card(
                    row + 1,
                    data["name"],
                    data["number"],
                    data["holder"],
                    data["expiry_month"],
                    data["expiry_year"],
                    data["cvv"],
                    data["type"],
                ):
                    QMessageBox.information(self, "Success", "Card updated!")
                    self._load_cards()
                else:
                    QMessageBox.warning(self, "Error", "Failed to update card.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _delete_selected_card(self):
        selected = self.cards_table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, "No Selection", "Please select a card to delete.")
            return

        name_item = self.cards_table.item(selected, 0)
        name = name_item.text() if name_item else "Unknown"

        reply = QMessageBox.question(
            self,
            "Delete Card",
            f"Are you sure you want to delete card '{name}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.controller.set_master_password(self.master_password)
                if self.controller.delete_card(selected + 1):
                    QMessageBox.information(self, "Success", "Card deleted!")
                    self._load_cards()
                else:
                    QMessageBox.warning(self, "Error", "Failed to delete card.")
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))

    def _add_new_note(self):
        if not self.master_password:
            QMessageBox.warning(self, "Locked", "Please unlock vault first!")
            return

        dialog = NoteDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_data()
            if not data["title"] or not data["content"]:
                QMessageBox.warning(self, "Error", "Title and content are required!")
                return

            try:
                self.controller.set_master_password(self.master_password)
                if self.controller.save_note(data["title"], data["content"]):
                    QMessageBox.information(self, "Success", "Note saved successfully!")
                    self._load_notes()
                else:
                    QMessageBox.warning(self, "Error", "Failed to save note.")
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))

    def _load_notes(self):
        if not self.master_password:
            return

        try:
            self.controller.set_master_password(self.master_password)
            notes = self.controller.get_notes()
            self.notes_table.setRowCount(len(notes))
            for row, note in enumerate(notes):
                self.notes_table.setItem(
                    row, 0, QTableWidgetItem(note.get("title", ""))
                )
                content = note.get("content", "")
                preview = content[:50] + ("..." if len(content) > 50 else "")
                self.notes_table.setItem(row, 1, QTableWidgetItem(preview))

                view_btn = QPushButton("👁️")
                view_btn.setFixedSize(60, 30)
                view_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #2a2a48;
                        color: #a0a0b8;
                        border: none;
                        border-radius: 6px;
                        font-size: 14px;
                        padding: 0px;
                        margin: 0px;
                        text-align: center;
                    }
                    QPushButton:hover {
                        background-color: #3a3a5a;
                        color: #f2c94c;
                    }
                """)
                view_btn.setToolTip("View note")
                view_btn.clicked.connect(
                    lambda checked, r=row: self._view_note_content(r)
                )
                self.notes_table.setCellWidget(row, 2, view_btn)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load notes: {str(e)}")

    def _view_note_content(self, row):
        if not self.master_password:
            return

        try:
            self.controller.set_master_password(self.master_password)
            notes = self.controller.get_notes()
            if row >= len(notes):
                return

            note = notes[row]
            title = note.get("title", "Untitled")
            content = note.get("content", "")
            created = note.get("created", "")

            dialog = QDialog(self)
            dialog.setWindowTitle(f"📝 {title}")
            dialog.setMinimumSize(460, 360)
            dialog.setStyleSheet("""
                QDialog {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #0f0f1a, stop:1 #1a1a2e);
                    border-radius: 16px;
                }
                QLabel {
                    color: #e8e8f0;
                    font-family: 'Segoe UI';
                }
                QTextEdit {
                    background-color: #14142a;
                    color: #e8e8f0;
                    border: 1px solid #2a2a48;
                    border-radius: 10px;
                    padding: 12px;
                    font-size: 13px;
                    font-family: 'Segoe UI';
                }
                QPushButton {
                    background-color: #4a6fa5;
                    color: white;
                    border: none;
                    border-radius: 10px;
                    padding: 8px 24px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #5a7fb5;
                }
            """)

            layout = QVBoxLayout(dialog)
            layout.setSpacing(12)

            title_label = QLabel(f"📌 {title}")
            title_label.setStyleSheet(
                "font-size: 18px; font-weight: bold; color: #f2c94c;"
            )
            layout.addWidget(title_label)

            content_display = QTextEdit()
            content_display.setPlainText(content)
            content_display.setReadOnly(True)
            layout.addWidget(content_display)

            info_label = QLabel(f"Created: {created}")
            info_label.setStyleSheet("color: #606080; font-size: 10px;")
            info_label.setAlignment(Qt.AlignmentFlag.AlignRight)
            layout.addWidget(info_label)

            close_btn = QPushButton("Close")
            close_btn.clicked.connect(dialog.accept)
            layout.addWidget(close_btn)

            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _edit_note(self, row, column):
        if not self.master_password:
            return

        try:
            self.controller.set_master_password(self.master_password)
            notes = self.controller.get_notes()
            if row >= len(notes):
                return

            note_data = notes[row]
            dialog = NoteDialog(self, note_data)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                data = dialog.get_data()
                if not data["title"] or not data["content"]:
                    QMessageBox.warning(
                        self, "Error", "Title and content are required!"
                    )
                    return

                if self.controller.update_note(row + 1, data["title"], data["content"]):
                    QMessageBox.information(self, "Success", "Note updated!")
                    self._load_notes()
                else:
                    QMessageBox.warning(self, "Error", "Failed to update note.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _delete_selected_note(self):
        selected = self.notes_table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, "No Selection", "Please select a note to delete.")
            return

        title_item = self.notes_table.item(selected, 0)
        title = title_item.text() if title_item else "Unknown"

        reply = QMessageBox.question(
            self,
            "Delete Note",
            f"Are you sure you want to delete note '{title}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.controller.set_master_password(self.master_password)
                if self.controller.delete_note(selected + 1):
                    QMessageBox.information(self, "Success", "Note deleted!")
                    self._load_notes()
                else:
                    QMessageBox.warning(self, "Error", "Failed to delete note.")
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))

    # Export functions (truncated for brevity, can be added later)
    def _export_passwords_txt(self):
        if not self.master_password:
            return
        try:
            self.controller.set_master_password(self.master_password)
            passwords = self.controller.get_passwords()
            if not passwords:
                QMessageBox.warning(self, "Empty", "No passwords to export.")
                return
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Passwords as TXT",
                f"passwords_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                "Text Files (*.txt)",
            )
            if not file_path:
                return
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("=" * 60 + "\n")
                f.write("SECUREPASS PRO - PASSWORDS EXPORT\n")
                f.write(f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("=" * 60 + "\n\n")
                for i, pwd in enumerate(passwords, 1):
                    f.write(f"{i:3}. Name: {pwd.get('name', '')}\n")
                    f.write(f"     Password: {pwd.get('password', '')}\n")
                    f.write("-" * 60 + "\n")
                f.write(f"\nTotal: {len(passwords)} passwords\n")
            QMessageBox.information(
                self, "Success", f"Passwords exported to:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export: {str(e)}")

    def _export_passwords_excel(self):
        if not self.master_password:
            return
        try:
            self.controller.set_master_password(self.master_password)
            passwords = self.controller.get_passwords()
            if not passwords:
                QMessageBox.warning(self, "Empty", "No passwords to export.")
                return
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Passwords as Excel",
                f"passwords_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)",
            )
            if not file_path:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "Passwords Export"
            ws["A1"] = "#"
            ws["B1"] = "Name"
            ws["C1"] = "Password"
            for i, pwd in enumerate(passwords, 1):
                ws[f"A{i+1}"] = i
                ws[f"B{i+1}"] = pwd.get("name", "")
                ws[f"C{i+1}"] = pwd.get("password", "")
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 40
            wb.save(file_path)
            QMessageBox.information(
                self, "Success", f"Passwords exported to:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export: {str(e)}")

    def _export_cards_txt(self):
        if not self.master_password:
            return
        try:
            self.controller.set_master_password(self.master_password)
            cards = self.controller.get_cards()
            if not cards:
                QMessageBox.warning(self, "Empty", "No cards to export.")
                return
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Cards as TXT",
                f"cards_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                "Text Files (*.txt)",
            )
            if not file_path:
                return
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("=" * 60 + "\n")
                f.write("SECUREPASS PRO - CARDS EXPORT\n")
                f.write(f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("=" * 60 + "\n\n")
                for i, card in enumerate(cards, 1):
                    f.write(f"{i:3}. Name: {card.get('name', '')}\n")
                    f.write(f"     Type: {card.get('type', '')}\n")
                    f.write(f"     Number: {card.get('number', '')}\n")
                    f.write(f"     Holder: {card.get('holder', '')}\n")
                    f.write(
                        f"     Expiry: {card.get('expiry_month', '')}/{card.get('expiry_year', '')}\n"
                    )
                    f.write(f"     CVV: {card.get('cvv', '')}\n")
                    f.write("-" * 60 + "\n")
                f.write(f"\nTotal: {len(cards)} cards\n")
            QMessageBox.information(self, "Success", f"Cards exported to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export: {str(e)}")

    def _export_cards_excel(self):
        if not self.master_password:
            return
        try:
            self.controller.set_master_password(self.master_password)
            cards = self.controller.get_cards()
            if not cards:
                QMessageBox.warning(self, "Empty", "No cards to export.")
                return
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Cards as Excel",
                f"cards_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)",
            )
            if not file_path:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "Cards Export"
            ws["A1"] = "#"
            ws["B1"] = "Name"
            ws["C1"] = "Type"
            ws["D1"] = "Number"
            ws["E1"] = "Holder"
            ws["F1"] = "Expiry"
            ws["G1"] = "CVV"
            for i, card in enumerate(cards, 1):
                ws[f"A{i+1}"] = i
                ws[f"B{i+1}"] = card.get("name", "")
                ws[f"C{i+1}"] = card.get("type", "")
                ws[f"D{i+1}"] = card.get("number", "")
                ws[f"E{i+1}"] = card.get("holder", "")
                ws[f"F{i+1}"] = (
                    f"{card.get('expiry_month', '')}/{card.get('expiry_year', '')}"
                )
                ws[f"G{i+1}"] = card.get("cvv", "")
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 25
            ws.column_dimensions["E"].width = 20
            ws.column_dimensions["F"].width = 12
            ws.column_dimensions["G"].width = 10
            wb.save(file_path)
            QMessageBox.information(self, "Success", f"Cards exported to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export: {str(e)}")

    def _export_notes_txt(self):
        if not self.master_password:
            return
        try:
            self.controller.set_master_password(self.master_password)
            notes = self.controller.get_notes()
            if not notes:
                QMessageBox.warning(self, "Empty", "No notes to export.")
                return
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Notes as TXT",
                f"notes_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                "Text Files (*.txt)",
            )
            if not file_path:
                return
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("=" * 60 + "\n")
                f.write("SECUREPASS PRO - NOTES EXPORT\n")
                f.write(f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("=" * 60 + "\n\n")
                for i, note in enumerate(notes, 1):
                    f.write(f"{i:3}. Title: {note.get('title', '')}\n")
                    f.write(f"     Content: {note.get('content', '')}\n")
                    f.write(f"     Created: {note.get('created', '')}\n")
                    f.write("-" * 60 + "\n")
                f.write(f"\nTotal: {len(notes)} notes\n")
            QMessageBox.information(self, "Success", f"Notes exported to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export: {str(e)}")

    def _export_notes_excel(self):
        if not self.master_password:
            return
        try:
            self.controller.set_master_password(self.master_password)
            notes = self.controller.get_notes()
            if not notes:
                QMessageBox.warning(self, "Empty", "No notes to export.")
                return
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Notes as Excel",
                f"notes_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)",
            )
            if not file_path:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "Notes Export"
            ws["A1"] = "#"
            ws["B1"] = "Title"
            ws["C1"] = "Content"
            ws["D1"] = "Created"
            for i, note in enumerate(notes, 1):
                ws[f"A{i+1}"] = i
                ws[f"B{i+1}"] = note.get("title", "")
                ws[f"C{i+1}"] = note.get("content", "")
                ws[f"D{i+1}"] = note.get("created", "")
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 50
            ws.column_dimensions["D"].width = 25
            wb.save(file_path)
            QMessageBox.information(self, "Success", f"Notes exported to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export: {str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setApplicationName("SecurePass Pro")

    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window, QColor(15, 15, 26))
    palette.setColor(QPalette.ColorRole.WindowText, QColor(232, 232, 240))
    palette.setColor(QPalette.ColorRole.Base, QColor(26, 26, 46))
    palette.setColor(QPalette.ColorRole.AlternateBase, QColor(20, 20, 36))
    palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(26, 26, 46))
    palette.setColor(QPalette.ColorRole.ToolTipText, QColor(232, 232, 240))
    palette.setColor(QPalette.ColorRole.Text, QColor(232, 232, 240))
    palette.setColor(QPalette.ColorRole.Button, QColor(42, 42, 64))
    palette.setColor(QPalette.ColorRole.ButtonText, QColor(232, 232, 240))
    palette.setColor(QPalette.ColorRole.BrightText, QColor(255, 255, 255))
    palette.setColor(QPalette.ColorRole.Highlight, QColor(74, 111, 165))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor(255, 255, 255))
    app.setPalette(palette)

    window = SecurePassGUI()
    window.show()
    sys.exit(app.exec())
